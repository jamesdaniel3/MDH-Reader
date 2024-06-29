import os
from datetime import datetime
from openpyxl import Workbook

def generate_instructor_feedback_workbook(info):
    """
    This function takes in the result of get_ta_feedback() with anonymous set to true and generates an Excel
    sheet containing all the information. Each TA will have their own page containing their comments.
    """
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create a new sheet for each TA and add their comments
    for ta_name, comments in info.items():
        sheet = wb.create_sheet(title=ta_name)
        for row_index, comment in enumerate(comments, start=1):
            sheet.cell(row=row_index, column=1, value=comment)

    # Save the workbook with today's date in the title
    today_date = datetime.today().strftime('%Y-%m-%d')
    file_name = f'TA Feedback {today_date}.xlsx'
    directory = "excel_files"

    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Save the file in the specified directory
    file_path = os.path.join(directory, file_name)
    wb.save(file_path)

    print("Sheet created")

